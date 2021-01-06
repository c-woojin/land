import re
from typing import List, Tuple, Dict
from datetime import date, timedelta
from math import floor

import openpyxl

from src.domain.entity.complex import Complex
from src.domain.values import Region


class LandXlsHandler:
    def __init__(self, file_name: str, data: List[Tuple[Region, List[Complex]]]):
        self.file_name = file_name
        self.data = data

    def write_raw_xls(self):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for region, complexes in self.data:
            current_sheet = wb.create_sheet()
            current_sheet.title = region.region_name
            current_sheet.append(['단지명', '입주연도', '총세대수', '공급(㎡)', '공급면적', '전용(㎡)', '전용면적', '세대', '매매최저가', '층',
                                  '전세최고가', '층', '매전갭', '전세가율', '매매평단', '전세평단', '방', '욕실', '구조', '타입', '비고', '거래일자',
                                  '대표평형'])
            for c in complexes:
                for p in c.pyeongs:
                    etc = ""
                    int_pyeong = p.int_pyeong
                    if p.trade_date and p.trade_date < date.today() - timedelta(days=365):
                        etc += f"최근 1년 내 매매거래 없었음."
                    elif p.trade_date is None:
                        etc += f"매매거래 없음."
                    if p.lease_date and p.lease_date < date.today() - timedelta(days=365):
                        etc += "\r" if len(etc) > 0 else etc
                        etc += f"최근 1년 내 전세거래 없었음."
                    elif p.lease_date is None:
                        etc += "\r" if len(etc) > 0 else etc
                        etc += f"전세거래 없음."
                    if (p.trade_date is None and p.lease_date is None) and p.is_restriction:
                        etc += "\r" if len(etc) > 0 else etc
                        etc += f"전매제한"

                    current_sheet.append(
                        [c.complex_name,
                         f"'{c.completion_date.strftime('%Y-%m')}",
                         c.total_household_count,
                         p.supply_area,
                         int_pyeong,
                         p.exclusive_area,
                         p.exclusive_pyeong,
                         p.house_hold_count,
                         p.low_trade_price if p.low_trade_price else "",
                         p.trade_floor if p.trade_floor else "",
                         p.high_lease_price if p.high_lease_price else "",
                         p.lease_floor if p.lease_floor else "",
                         p.low_trade_price - p.high_lease_price if p.low_trade_price and p.high_lease_price else '',
                         str(floor(p.high_lease_price / p.low_trade_price * 100)) + '%' if p.low_trade_price and p.high_lease_price else '',
                         '@' + str(round(p.low_trade_price / int_pyeong)) if p.low_trade_price else 0,
                         '@' + str(round(p.high_lease_price / int_pyeong)) if p.high_lease_price else 0,
                         p.room_count, p.bathroom_count, p.formatting_entrance_type, p.pyeong_name, etc,
                         (p.trade_date.strftime("매매 : %y/%m\r") if p.trade_date else "") + (
                            p.lease_date.strftime("전세 : %y/%m") if p.lease_date else ""),
                         "v" if p.is_representative else ""]
                    )
        wb.save(self.file_name)

    def write_analysis_xls(self, latest_year: int, sub_latest_year: int):
        wb = openpyxl.Workbook()
        sheet = wb.active

        prices_by_towns_10 = {town: {'old': [], 'sub_latest': [], 'latest': []} for town, complexes in self.data}
        high_price_10 = 0
        low_price_10 = 9999999999999
        prices_by_towns_20 = {town: {'old': [], 'sub_latest': [], 'latest': []} for town, complexes in self.data}
        high_price_20 = 0
        low_price_20 = 9999999999999
        prices_by_towns_20_2 = {town: {'old': [], 'sub_latest': [], 'latest': []} for town, complexes in self.data}
        high_price_20_2 = 0
        low_price_20_2 = 9999999999999
        prices_by_towns_20_3 = {town: {'old': [], 'sub_latest': [], 'latest': []} for town, complexes in self.data}
        high_price_20_3 = 0
        low_price_20_3 = 9999999999999
        prices_by_towns_30 = {town: {'old': [], 'sub_latest': [], 'latest': []} for town, complexes in self.data}
        high_price_30 = 0
        low_price_30 = 9999999999999
        prices_by_towns_40 = {town: {'old': [], 'sub_latest': [], 'latest': []} for town, complexes in self.data}
        high_price_40 = 0
        low_price_40 = 9999999999999

        for town, complexes in self.data:
            for c in complexes:
                for p in c.pyeongs:
                    if p.is_representative and (10 <= p.int_pyeong < 50):
                        if p.low_trade_price:
                            # high_price setting
                            if p.int_pyeong < 20 and p.low_trade_price > high_price_10:
                                high_price_10 = p.low_trade_price
                            elif p.int_pyeong < 30 and p.low_trade_price > high_price_20:
                                high_price_20 = p.low_trade_price
                                if p.room_count == 2 and p.low_trade_price > high_price_20_2:
                                    high_price_20_2 = p.low_trade_price
                                elif p.room_count == 3 and p.low_trade_price > high_price_20_3:
                                    high_price_20_3 = p.low_trade_price
                            elif p.int_pyeong < 40 and p.low_trade_price > high_price_30:
                                high_price_30 = p.low_trade_price
                            elif p.low_trade_price > high_price_40:
                                high_price_40 = p.low_trade_price

                            # low_price setting
                            if p.int_pyeong < 20 and p.low_trade_price < low_price_10:
                                low_price_10 = p.low_trade_price
                            elif p.int_pyeong < 30 and p.low_trade_price < low_price_20:
                                low_price_20 = p.low_trade_price
                                if p.room_count == 2 and p.low_trade_price < low_price_20_2:
                                    low_price_20_2 = p.low_trade_price
                                elif p.room_count == 3 and p.low_trade_price < low_price_20_3:
                                    low_price_20_3 = p.low_trade_price
                            elif p.int_pyeong < 40 and p.low_trade_price < low_price_30:
                                low_price_30 = p.low_trade_price
                            elif p.low_trade_price < low_price_40:
                                low_price_40 = p.low_trade_price

                            # decision to latest apt
                            if c.completion_date.year >= latest_year:
                                key = "latest"
                            elif c.completion_date.year >= sub_latest_year:
                                key = "sub_latest"
                            else:
                                key = "old"

                            # categorizes
                            if p.int_pyeong < 20:
                                self.set_prices_by_towns(prices_by_towns_10, town, key, c, p)
                            elif p.int_pyeong < 30:
                                self.set_prices_by_towns(prices_by_towns_20, town, key, c, p)
                                if p.room_count == 2:
                                    self.set_prices_by_towns(prices_by_towns_20_2, town, key, c, p)
                                elif p.room_count == 3:
                                    self.set_prices_by_towns(prices_by_towns_20_3, town, key, c, p)
                            elif p.int_pyeong < 40:
                                self.set_prices_by_towns(prices_by_towns_30, town, key, c, p)
                            else:
                                self.set_prices_by_towns(prices_by_towns_40, town, key, c, p)

        rows_10 = self._generate_rows_prices_by_towns(prices_by_towns_10, high_price_10, low_price_10)
        rows_20 = self._generate_rows_prices_by_towns(prices_by_towns_20, high_price_20, low_price_20)
        rows_20_2 = self._generate_rows_prices_by_towns(prices_by_towns_20_2, high_price_20_2, low_price_20_2)
        rows_20_3 = self._generate_rows_prices_by_towns(prices_by_towns_20_3, high_price_20_3, low_price_20_3)
        rows_30 = self._generate_rows_prices_by_towns(prices_by_towns_30, high_price_30, low_price_30)
        rows_40 = self._generate_rows_prices_by_towns(prices_by_towns_40, high_price_40, low_price_40)

        sheet.title = "10평"
        self._write_prices_by_towns_xls(sheet, prices_by_towns_10, latest_year, sub_latest_year, rows_10)
        sheet = wb.create_sheet("20평")
        self._write_prices_by_towns_xls(sheet, prices_by_towns_20, latest_year, sub_latest_year, rows_20)
        sheet = wb.create_sheet("20평(방2개)")
        self._write_prices_by_towns_xls(sheet, prices_by_towns_20_2, latest_year, sub_latest_year, rows_20_2)
        sheet = wb.create_sheet("20평(방3개)")
        self._write_prices_by_towns_xls(sheet, prices_by_towns_20_3, latest_year, sub_latest_year, rows_20_3)
        sheet = wb.create_sheet("30평")
        self._write_prices_by_towns_xls(sheet, prices_by_towns_30, latest_year, sub_latest_year, rows_30)
        sheet = wb.create_sheet("40평")
        self._write_prices_by_towns_xls(sheet, prices_by_towns_40, latest_year, sub_latest_year, rows_40)

        wb.save(self.file_name)

    @staticmethod
    def set_prices_by_towns(prices_by_towns, town, key, c, p):
        prices_by_towns.get(town).get(key).append({
            'name': c.complex_name,
            'pyeong': p.int_pyeong,
            'price': float(p.low_trade_price / 10000),
            'approved_year': c.completion_date.year
        })

    @staticmethod
    def _generate_rows_prices_by_towns(prices_by_towns: Dict[Region, Dict[str, List[Dict]]], high_price, low_price):
        # initialize
        gap = float(round((high_price - low_price) / 90000, 1))
        start_price = float(floor(low_price / 10000))
        result = []
        for i in range(10):
            result.append(
                {'period': (start_price, round(start_price + gap, 1)),
                 'complexes': {(town.region_name, new_key): [] for town, complex_dict in prices_by_towns.items() for new_key in complex_dict.keys()}
                 }
            )
            start_price = round(start_price + gap, 1)
        result.reverse()

        # set prices by period
        for town, complex_dict in prices_by_towns.items():
            for new_key, complexes in complex_dict.items():
                complexes.sort(key=lambda x: x.get('price'), reverse=True)
                for d in complexes:
                    for r in result:
                        low, high = r.get('period')
                        if low <= d.get('price') < high:
                            r.get('complexes').get((town.region_name, new_key)).append(f"{d.get('name')}({d.get('pyeong')}평/{d.get('price')}/{str(d.get('approved_year'))[2:]}년)")

        return result

    @staticmethod
    def _write_prices_by_towns_xls(sheet, prices_by_towns: Dict[Region, Dict[str, List[Dict]]], latest_year,
                                   sub_latest_year, result):
        sheet.append(['동네'] + [town.region_name for town, cd in prices_by_towns.items() for _ in cd.keys()])
        headers = ["시기"]
        for _ in prices_by_towns.keys():
            headers.append(f"구축(~{sub_latest_year})")
            headers.append(f"준신축({sub_latest_year}~{latest_year})")
            headers.append(f"신축({latest_year}~)")
        sheet.append(headers)
        for r in result:
            complexes = r.get('complexes')
            row = [f"{r.get('period')[0]} ~ {r.get('period')[1]}"]
            for town, complex_dict in prices_by_towns.items():
                c = ""
                for new_key in complex_dict.keys():
                    for cx in complexes.get((town.region_name, new_key)):
                        c = c + "\r" if len(c) > 0 else c
                        c += cx
                    row.append(f"{c}")
            sheet.append(row)
